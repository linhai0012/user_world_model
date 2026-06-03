"""Merge 4-persona Phase 2 user_gen_judge outputs into one xlsx for review.

Input: dynamic_usersim/outputs/user_gen_pid{0,4,12,14}_step200.json
Output: dynamic_usersim/outputs/user_gen_phase2_step200_review.xlsx

Each persona gets its own sheet. Within a sheet, one row per sample (50 per
persona); columns = chatbot_prev_snippet + gt + per-condition (gen, score).

Conditions in order: base_demo, base_full, teacher_demo, teacher_full,
student_demo. Colour coding:
  - header row bold
  - student_demo columns highlighted blue (the main thing we care about)
  - score columns: red for 1, orange for 2, yellow for 3, lightgreen for 4, green for 5

Usage:
    python dynamic_usersim/student_opd/user_gen_to_xlsx.py
"""
from __future__ import annotations
import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


CONDITIONS = ["base_demo", "base_full", "teacher_demo",
              "teacher_full", "student_demo"]

STUDENT_FILL = PatternFill("solid", fgColor="E3F2FD")  # light blue
HEADER_FILL = PatternFill("solid", fgColor="ECEFF1")
SCORE_FILLS = {
    1: PatternFill("solid", fgColor="FFCDD2"),
    2: PatternFill("solid", fgColor="FFE0B2"),
    3: PatternFill("solid", fgColor="FFF9C4"),
    4: PatternFill("solid", fgColor="DCEDC8"),
    5: PatternFill("solid", fgColor="A5D6A7"),
}


def build_persona_sheet(wb: Workbook, pid: int, data: dict) -> None:
    ws = wb.create_sheet(f"pid{pid}")
    pc = data["per_condition"]
    # align samples by sample_idx (they should be parallel across conditions)
    all_idxs = sorted(set(e["sample_idx"] for c in CONDITIONS
                          for e in pc.get(c, [])))

    # Build header
    header = ["idx", "chatbot_prev_snippet", "gt"]
    for cond in CONDITIONS:
        header += [f"{cond}_gen", f"{cond}_score"]
    for ci, h in enumerate(header, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True)
        c.fill = HEADER_FILL

    col_widths = {1: 5, 2: 45, 3: 55}
    for ci in range(4, len(header) + 1):
        col_widths[ci] = 55 if "gen" in header[ci - 1] else 8
    for ci, w in col_widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Index entries by (cond, sample_idx) for fast lookup
    idxed = {
        cond: {e["sample_idx"]: e for e in pc.get(cond, [])}
        for cond in CONDITIONS
    }

    for ri, sidx in enumerate(all_idxs, start=2):
        ws.cell(row=ri, column=1, value=sidx)
        # pull chatbot_prev_snippet + gt from any condition (they should match)
        ref = None
        for cond in CONDITIONS:
            if sidx in idxed[cond]:
                ref = idxed[cond][sidx]
                break
        if ref is None:
            continue
        ws.cell(row=ri, column=2, value=ref.get("chatbot_prev_snippet", ""))
        ws.cell(row=ri, column=3, value=ref.get("gt", ""))

        col = 4
        for cond in CONDITIONS:
            e = idxed[cond].get(sidx)
            if e is None:
                col += 2
                continue
            gen_cell = ws.cell(row=ri, column=col, value=e.get("gen", ""))
            score = e.get("score")
            score_cell = ws.cell(row=ri, column=col + 1, value=score)
            # colouring
            if cond == "student_demo":
                gen_cell.fill = STUDENT_FILL
            if score in SCORE_FILLS:
                score_cell.fill = SCORE_FILLS[score]
            col += 2

        # wrap long text
        for ci in range(1, len(header) + 1):
            ws.cell(row=ri, column=ci).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

    ws.freeze_panes = "D2"  # freeze idx + snippet + gt


def build_summary_sheet(wb: Workbook, per_pid: dict[int, dict]) -> None:
    ws = wb.create_sheet("summary", 0)
    header = ["pid"] + [f"{c}_mean" for c in CONDITIONS] + ["n"]
    for ci, h in enumerate(header, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True)
    row = 2
    for pid in sorted(per_pid):
        data = per_pid[pid]
        means = []
        n = 0
        for cond in CONDITIONS:
            entries = data["per_condition"].get(cond, [])
            scores = [e["score"] for e in entries
                      if isinstance(e.get("score"), (int, float))]
            mean = sum(scores) / len(scores) if scores else None
            means.append(mean)
            n = max(n, len(entries))
        ws.cell(row=row, column=1, value=pid)
        for i, m in enumerate(means, start=2):
            cell = ws.cell(row=row, column=i,
                           value=round(m, 3) if m is not None else "")
        ws.cell(row=row, column=len(CONDITIONS) + 2, value=n)
        row += 1
    for ci, w in enumerate([5] + [14] * len(CONDITIONS) + [5], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--in-dir", type=Path,
                    default=Path("dynamic_usersim/outputs"))
    ap.add_argument("--pids", type=int, nargs="+",
                    default=[0, 4, 12, 14])
    ap.add_argument("--step", type=str, default="step200")
    ap.add_argument("--out", type=Path,
                    default=Path("dynamic_usersim/outputs")
                    / "user_gen_phase2_step200_review.xlsx")
    args = ap.parse_args()

    per_pid = {}
    for pid in args.pids:
        path = args.in_dir / f"user_gen_pid{pid}_{args.step}.json"
        if not path.exists():
            print(f"MISSING: {path}")
            continue
        with path.open(encoding="utf-8") as f:
            per_pid[pid] = json.load(f)
        print(f"loaded {path}  "
              f"({per_pid[pid]['n_samples']} samples, "
              f"{len(per_pid[pid]['per_condition'])} conditions)")

    if not per_pid:
        print("no inputs found")
        return

    wb = Workbook()
    # remove default empty sheet (we create one per persona)
    wb.remove(wb.active)

    build_summary_sheet(wb, per_pid)
    for pid, data in sorted(per_pid.items()):
        build_persona_sheet(wb, pid, data)

    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"wrote {args.out}")


if __name__ == "__main__":
    main()
