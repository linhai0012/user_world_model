"""Merge verbal_pid{N}_base.jsonl + verbal_pid{N}_r1b.jsonl into a single
side-by-side xlsx for human review.

Each MCQ → 4 rows (one per choice). Columns let you compare at a glance:
  - choice_text
  - reaction_base  (UserSim reaction from base Qwen3 Instruct)
  - reaction_r1b   (UserSim reaction from R1b LoRA merged)
  - gt_asst        (real assistant turn at end_index+1)
  - gt_user        (real user followup at end_index+2 — our "gold" reference)

Correct-choice rows highlighted green. Qtype grouped by inserting blank row
between MCQs for readability.

Usage:
    python dynamic_usersim/student_opd/verbal_to_xlsx.py \\
        --base dynamic_usersim/outputs/verbal_pid4_base.jsonl \\
        --r1b  dynamic_usersim/outputs/verbal_pid4_r1b.jsonl \\
        --out  dynamic_usersim/outputs/verbal_pid4_review.xlsx
"""
from __future__ import annotations
import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


COLS = [
    ("qid_short",        12),
    ("qtype",            22),
    ("topic",            22),
    ("user_question",    40),
    ("label",             5),
    ("correct",           7),
    ("choice_text",      60),
    ("reaction_base",    55),
    ("reaction_r1b",     55),
    ("reaction_phase2",  55),
    ("gt_asst",          45),
    ("gt_user",          45),
]


def _first(lst, idx=0):
    try:
        return lst[idx]
    except (IndexError, TypeError):
        return ""


def load_jsonl(path: Path) -> dict[str, dict]:
    """Return qid → record dict."""
    out = {}
    with path.open(encoding="utf-8") as f:
        for line in f:
            r = json.loads(line)
            out[r["question_id"]] = r
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--base", type=Path, required=True)
    ap.add_argument("--r1b",  type=Path, required=True)
    ap.add_argument("--phase2", type=Path, default=None,
                    help="optional phase2 single-LoRA output for 3-way compare")
    ap.add_argument("--out",  type=Path, required=True)
    args = ap.parse_args()

    base = load_jsonl(args.base)
    r1b  = load_jsonl(args.r1b)
    p2   = load_jsonl(args.phase2) if args.phase2 else {}

    qids = [q for q in base.keys() if q in r1b]
    if p2:
        qids = [q for q in qids if q in p2]
    print(f"common MCQs: {len(qids)}  "
          f"(base={len(base)}, r1b={len(r1b)}, phase2={len(p2)})")

    # Sort by qtype then qid so same types group together
    qids.sort(key=lambda q: (base[q]["qtype_canonical"], q))

    wb = Workbook()
    ws = wb.active
    ws.title = "reactions"

    # Header
    hdr_font = Font(bold=True)
    for ci, (name, width) in enumerate(COLS, start=1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font = hdr_font
        ws.column_dimensions[get_column_letter(ci)].width = width

    correct_fill = PatternFill("solid", fgColor="E8F5E9")   # light green
    qtype_break_fill = PatternFill("solid", fgColor="F5F5F5")  # light gray

    row = 2
    prev_qtype = None
    for qid in qids:
        br = base[qid]
        rr = r1b[qid]
        if br["qtype_canonical"] != prev_qtype and prev_qtype is not None:
            # blank separator row (so eye can group)
            ws.row_dimensions[row].height = 6
            row += 1
        prev_qtype = br["qtype_canonical"]

        gt = br.get("gt_followup", [])
        gt_asst = _first(gt, 0).get("content", "") if gt else ""
        gt_user = _first(gt, 1).get("content", "") if len(gt) > 1 else ""

        # map label → base and r1b reaction
        br_react = {c["label"]: _first(c["reactions"]) for c in br["choices"]}
        rr_react = {c["label"]: _first(c["reactions"]) for c in rr["choices"]}
        p2_react = ({c["label"]: _first(c["reactions"]) for c in p2[qid]["choices"]}
                    if p2 else {})
        label_choice = {c["label"]: c for c in br["choices"]}

        for label in ("a", "b", "c", "d"):
            if label not in label_choice:
                continue
            c = label_choice[label]
            vals = {
                "qid_short":       qid[:8],
                "qtype":           br["qtype_canonical"],
                "topic":           br["topic"],
                "user_question":   br["user_question"],
                "label":           label,
                "correct":         "✓" if c["is_correct"] else "",
                "choice_text":     c["choice_text"],
                "reaction_base":   br_react.get(label, ""),
                "reaction_r1b":    rr_react.get(label, ""),
                "reaction_phase2": p2_react.get(label, ""),
                "gt_asst":         gt_asst,
                "gt_user":         gt_user,
            }
            for ci, (name, _) in enumerate(COLS, start=1):
                cell = ws.cell(row=row, column=ci, value=vals[name])
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if c["is_correct"]:
                    cell.fill = correct_fill
            row += 1

    # Freeze header
    ws.freeze_panes = "A2"

    # --- Sheet 2: summary per qtype ---
    ws2 = wb.create_sheet("summary")
    ws2.append(["qtype", "n_mcqs", "notes"])
    from collections import Counter
    ct = Counter(base[q]["qtype_canonical"] for q in qids)
    for qt, n in sorted(ct.items()):
        ws2.append([qt, n, ""])
    for ci, w in enumerate([24, 10, 40], start=1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"wrote {args.out}  ({row-1} data rows)")


if __name__ == "__main__":
    main()
