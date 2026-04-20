"""6-way comparison xlsx across all verbal Stage-1 output jsonls for pid=4.

Columns per (MCQ, choice) row:
  - qid_short, qtype, topic
  - user_question, label, correct, choice_text
  - reaction_base_hf / r1b_hf / phase2_hf      (broken HF pipeline — 87% identical)
  - reaction_base_vllm / phase2_vllm / r1b_vllm  (fixed vLLM pipeline)
  - gt_asst, gt_user

Correct-choice rows highlighted green. Qtype groups separated by thin blank row.
"""
from __future__ import annotations
import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


COLS = [
    ("qid_short",            12),
    ("qtype",                22),
    ("user_question",        40),
    ("label",                 5),
    ("correct",               7),
    ("choice_text",          55),
    ("reaction_base_vllm",   40),
    ("reaction_phase2_vllm", 40),
    ("reaction_r1b_vllm",    40),
    ("reaction_opsd",        40),
    ("gt_asst",              35),
    ("gt_user",              35),
]

CORRECT_FILL = PatternFill("solid", fgColor="E8F5E9")
VLLM_FILL = PatternFill("solid", fgColor="E3F2FD")


def load_jsonl(path: Path) -> dict[str, dict]:
    out = {}
    with path.open(encoding="utf-8") as f:
        for line in f:
            r = json.loads(line)
            out[r["question_id"]] = r
    return out


def first_reaction(rec, label):
    for c in rec.get("choices", []):
        if c["label"] == label:
            return (c["reactions"][0] if c["reactions"] else "")
    return ""


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--in-dir", type=Path,
                    default=Path("dynamic_usersim/outputs"))
    ap.add_argument("--out", type=Path,
                    default=Path("dynamic_usersim/outputs")
                    / "verbal_pid4_review_6way.xlsx")
    args = ap.parse_args()

    files = {
        "base_vllm":     args.in_dir / "verbal_pid4_base_vllm.jsonl",
        "phase2_vllm":   args.in_dir / "verbal_pid4_phase2_vllm.jsonl",
        "r1b_vllm":      args.in_dir / "verbal_pid4_r1b_vllm.jsonl",
        "opsd":          args.in_dir / "verbal_pid4_opsd.jsonl",
    }
    data = {k: load_jsonl(p) for k, p in files.items() if p.exists()}
    if not data:
        raise SystemExit("no input jsonls found")

    qids_all = sorted(set.intersection(*(set(d.keys()) for d in data.values())))

    # group by qtype
    by_qtype: dict[str, list[str]] = {}
    for qid in qids_all:
        qt = data["base_vllm"][qid]["qtype_canonical"]
        by_qtype.setdefault(qt, []).append(qid)
    for qt in by_qtype:
        by_qtype[qt].sort()

    wb = Workbook()
    wb.remove(wb.active)
    total_rows = 0

    for qtype in sorted(by_qtype):
        ws = wb.create_sheet(qtype[:31])  # xlsx sheet name max 31 chars
        for ci, (name, width) in enumerate(COLS, 1):
            c = ws.cell(row=1, column=ci, value=name)
            c.font = Font(bold=True)
            if name.endswith("_vllm"):
                c.fill = VLLM_FILL
            ws.column_dimensions[get_column_letter(ci)].width = width

        row = 2
        for qid in by_qtype[qtype]:
            ref = data["base_vllm"][qid]
            gt = ref.get("gt_followup", [])
            gt_asst = gt[0].get("content", "") if gt else ""
            gt_user = gt[1].get("content", "") if len(gt) > 1 else ""
            label_to_choice = {c["label"]: c for c in ref["choices"]}
            correct_label = ref["correct_answer"]

            for label in ("a", "b", "c", "d"):
                if label not in label_to_choice:
                    continue
                c = label_to_choice[label]
                is_correct = (label == correct_label)
                vals = {
                    "qid_short":            qid[:8],
                    "qtype":                ref["qtype_canonical"],
                    "user_question":        ref["user_question"],
                    "label":                label,
                    "correct":              "✓" if is_correct else "",
                    "choice_text":          c["choice_text"],
                    "reaction_base_vllm":   first_reaction(data.get("base_vllm", {}).get(qid, {}), label),
                    "reaction_phase2_vllm": first_reaction(data.get("phase2_vllm", {}).get(qid, {}), label),
                    "reaction_r1b_vllm":    first_reaction(data.get("r1b_vllm", {}).get(qid, {}), label),
                    "reaction_opsd":        first_reaction(data.get("opsd", {}).get(qid, {}), label),
                    "gt_asst":              gt_asst,
                    "gt_user":              gt_user,
                }
                for ci, (name, _) in enumerate(COLS, 1):
                    cell = ws.cell(row=row, column=ci, value=vals[name])
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    if is_correct:
                        cell.fill = CORRECT_FILL
                    elif name.endswith("_vllm"):
                        cell.fill = VLLM_FILL
                row += 1
            # blank row between MCQs inside a qtype sheet
            ws.row_dimensions[row].height = 6
            row += 1

        ws.freeze_panes = "D2"
        print(f"  sheet {qtype:<22} {len(by_qtype[qtype]):3d} MCQs  {row-2} rows")
        total_rows += row - 2

    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"wrote {args.out}  ({total_rows} total rows, "
          f"{len(qids_all)} MCQs across {len(by_qtype)} qtypes)")


if __name__ == "__main__":
    main()
