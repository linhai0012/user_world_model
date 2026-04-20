"""Build a per-qtype xlsx from judge_verbal_golden_pid4.json for manual review.

Each qtype gets a sheet. Each MCQ is one row with 4 config columns:
  qid_short | qtype | topic | user_question | correct_choice_text
  | golden_snippet | reaction_base + score | reaction_phase2 + score
  | reaction_r1b + score | reaction_opsd + score

Scores colored: 1=red, 2=orange, 3=yellow, 4=lightgreen, 5=green.
Highest-score config per row is bolded.
"""
from __future__ import annotations
import argparse
import json
from pathlib import Path
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


CONFIG_ORDER = ["base_vllm", "phase2_vllm", "r1b_vllm", "opsd"]

COLS = [
    ("qid",                10),
    ("qtype",               20),
    ("topic",               20),
    ("user_question",       38),
    ("correct_choice_text", 50),
    ("golden_snippet",      50),
]
# 8 more columns = reaction + score per config × 4
for cfg in CONFIG_ORDER:
    COLS.append((f"reaction_{cfg}", 42))
    COLS.append((f"score_{cfg}",    7))


HEADER_FILL = PatternFill("solid", fgColor="ECEFF1")
SCORE_FILLS = {
    1: PatternFill("solid", fgColor="FFCDD2"),
    2: PatternFill("solid", fgColor="FFE0B2"),
    3: PatternFill("solid", fgColor="FFF9C4"),
    4: PatternFill("solid", fgColor="DCEDC8"),
    5: PatternFill("solid", fgColor="A5D6A7"),
}


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--in-json", type=Path,
                    default=Path("dynamic_usersim/outputs/judge_verbal_golden_pid4.json"))
    ap.add_argument("--out-xlsx", type=Path,
                    default=Path("dynamic_usersim/outputs/judge_verbal_golden_pid4.xlsx"))
    args = ap.parse_args()

    data = json.loads(args.in_json.read_text(encoding="utf-8"))
    results = data["results"]

    # Group by (qid), collect per-config cells
    per_qid: dict[str, dict] = defaultdict(dict)
    for r in results:
        qid = r["qid"]
        cfg = r["config"]
        per_qid[qid][cfg] = r
        per_qid[qid].setdefault("_meta", {
            "qid":   qid,
            "qtype": r.get("qtype", ""),
            "topic": r.get("topic", "") if "topic" in r else "",
            "choice_text": r.get("choice_text", ""),
            "gt": r.get("gt", ""),
            "correct_label": r.get("correct_label", ""),
        })
    # user_question not in results json — re-load via source jsonls?
    # Actually it IS in the jsonls but we didn't carry over. Let's re-load from
    # one of them:
    # Adopt "user_question from base_vllm jsonl by qid" strategy.
    try:
        with open("dynamic_usersim/outputs/verbal_pid4_base_vllm.jsonl",
                  encoding="utf-8") as f:
            for line in f:
                rec = json.loads(line)
                qid = rec["question_id"]
                if qid in per_qid:
                    per_qid[qid]["_meta"]["user_question"] = rec.get("user_question", "")
    except FileNotFoundError:
        pass

    # Group qids by qtype
    by_qtype: dict[str, list[str]] = defaultdict(list)
    for qid, cells in per_qid.items():
        qt = cells["_meta"]["qtype"]
        by_qtype[qt].append(qid)

    wb = Workbook()
    wb.remove(wb.active)

    # Summary first
    ws0 = wb.create_sheet("summary")
    s = data["summary"]
    ws0.append(["config"] + ["mean", "n", "1s", "2s", "3s", "4s", "5s", "parse_fail"])
    for cfg in CONFIG_ORDER:
        d = s.get(cfg, {})
        dist = d.get("distribution", [0, 0, 0, 0, 0])
        ws0.append([cfg, round(d.get("mean") or 0.0, 3),
                    d.get("n", 0), *dist, d.get("parse_fail", 0)])
    for ci, w in enumerate([16, 10, 8, 6, 6, 6, 6, 6, 12], 1):
        ws0.column_dimensions[get_column_letter(ci)].width = w
    for row in ws0.iter_rows(min_row=1, max_row=1):
        for c in row:
            c.font = Font(bold=True)
            c.fill = HEADER_FILL

    # One sheet per qtype
    for qtype in sorted(by_qtype):
        ws = wb.create_sheet(qtype[:31])
        # header
        for ci, (name, width) in enumerate(COLS, 1):
            c = ws.cell(row=1, column=ci, value=name)
            c.font = Font(bold=True)
            c.fill = HEADER_FILL
            ws.column_dimensions[get_column_letter(ci)].width = width

        row = 2
        for qid in sorted(by_qtype[qtype]):
            cells = per_qid[qid]
            meta = cells["_meta"]
            # compute best config for highlighting
            scores = {cfg: cells.get(cfg, {}).get("score") for cfg in CONFIG_ORDER}
            valid_scores = [(cfg, s) for cfg, s in scores.items() if isinstance(s, int)]
            best_score = max((s for _, s in valid_scores), default=None)
            best_cfgs = {cfg for cfg, s in valid_scores if s == best_score} if best_score is not None else set()

            vals = {
                "qid":                 qid[:8],
                "qtype":               meta["qtype"],
                "topic":               meta.get("topic", ""),
                "user_question":       meta.get("user_question", ""),
                "correct_choice_text": meta["choice_text"],
                "golden_snippet":      meta["gt"],
            }
            for cfg in CONFIG_ORDER:
                rec = cells.get(cfg, {})
                vals[f"reaction_{cfg}"] = rec.get("gen", "")
                vals[f"score_{cfg}"] = rec.get("score", "")

            for ci, (name, _) in enumerate(COLS, 1):
                cell = ws.cell(row=row, column=ci, value=vals[name])
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                # Color score cells
                if name.startswith("score_"):
                    s = vals[name]
                    if isinstance(s, int) and s in SCORE_FILLS:
                        cell.fill = SCORE_FILLS[s]
                # Bold the best-config columns
                if name.startswith("reaction_") or name.startswith("score_"):
                    cfg = name.split("_", 1)[1]
                    if cfg in best_cfgs and best_score is not None and best_score >= 2:
                        cell.font = Font(bold=True)
            row += 1

        ws.freeze_panes = "G2"

    args.out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out_xlsx)
    print(f"wrote {args.out_xlsx}  ({sum(len(v) for v in by_qtype.values())} MCQs "
          f"across {len(by_qtype)} qtypes)")


if __name__ == "__main__":
    main()
